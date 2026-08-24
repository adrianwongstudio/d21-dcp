"""Build docs/inyear.xlsx - the in-year workbook area directors take into a
conversation with club officers.

One row per club: what the dashboard says today, what is still reachable, the
membership position, and the twelve underlying goal counts so an officer can
see exactly which number is short and by how much. Reads docs/live.json, so
run gen_live_data.py first.
"""
import os,json,collections,datetime
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

_ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def _p(*a): return os.path.join(_ROOT,*a)

TARGETS=[4,2,2,2,1,1,4,4,4,4,1,1]
ROWLBL=["Level 1 awards","Level 2 awards","More Level 2 awards","Level 3 awards",
 "Level 4 / PC / DTM","A second Level 4 / PC / DTM","New members","More new members",
 "Officers trained Jun-Aug","Officers trained Nov-Feb","Renewal dues on time","Officer list on time"]

HDRF=Font(bold=True,color="FFFFFF",size=10); HDRB=PatternFill("solid",fgColor="1F3864")
SUBB=PatternFill("solid",fgColor="2E5A94")
MET =PatternFill("solid",fgColor="D6EBDA"); SHORT=PatternFill("solid",fgColor="FBE0DD")
SHUT=PatternFill("solid",fgColor="E4E9ED")
BOLD=Font(bold=True); MUT=Font(color="6E828F",size=9)
THIN=Border(bottom=Side(style='thin',color="D7DEE4"))

def prior_year():
    """Last closed year's final score, for context in the conversation."""
    f=_p('data','rows.json')
    if not os.path.exists(f): return {},''
    rows=json.load(open(f))
    pys=sorted({r['Program Year'] for r in rows})
    if not pys: return {},''
    last=pys[-1]; out={}
    for r in rows:
        if r['Program Year']==last and r['Month End'][5:7]=='06':
            out[r['Club No']]=(r.get('DCP Goals Met'),r.get('DCP Status') or '')
    return out,last

def main():
    L=json.load(open(_p('docs','live.json')))
    prior,pyl=prior_year()
    wb=Workbook(); ws=wb.active; ws.title="Club Health"

    head=["Division","Area","Club No","Club Name","Goals met","of 10","Ceiling",
          "Best still possible","Members","Base","Net growth","Meets membership rule",
          "Next deadline","What closes then","Club Success Plan"]
    head+= [f"{ROWLBL[i]} (need {TARGETS[i]})" for i in range(12)]
    head+= [f"{pyl} final",f"{pyl} status"] if pyl else []
    ws.append(head)
    for c in range(1,len(head)+1):
        cell=ws.cell(row=1,column=c); cell.font=HDRF
        cell.fill=HDRB if c<=15 else SUBB
        cell.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center')
    ws.row_dimensions[1].height=46

    clubs=sorted(L['clubs'],key=lambda c:(c['d'] or 'zz',c['a'] or 'zz',c['m'].lower()))
    for c in clubs:
        pf,pst=prior.get(c['n'],(None,''))
        row=[c['d'],c['a'],int(c['n']),c['m'],c['met'],10,c['ceil'],c['best'] or "none",
             c['md'],c['mb'],c['ng'],"yes" if c['memok'] else "no",
             (datetime.date.fromisoformat(c['nd']) if c['nd'] else None),c['ndl'] or "",
             ("submitted" if (c.get("csp") or "").lower().startswith("requirement met") else
              "not yet" if c.get("csp") else "")]
        row+= [(c['v'][i] if c['v'][i] is not None else "") for i in range(12)]
        if pyl: row+=[pf,pst]
        ws.append(row)
        r=ws.max_row
        # shade each goal cell by whether it is met, and grey out shut windows
        for i in range(12):
            cell=ws.cell(row=r,column=16+i)
            cell.alignment=Alignment(horizontal='center')
            v=c['v'][i]
            # rows 8+9 share goal 9, rows 10+11 share goal 10
            state=c['st'][i] if i<8 else (c['st'][8] if i in (8,9) else c['st'][9])
            if v is not None and v>=TARGETS[i]: cell.fill=MET
            elif state=='d': cell.fill=SHUT
            else: cell.fill=SHORT
        ws.cell(row=r,column=5).font=BOLD
        ws.cell(row=r,column=6).font=MUT
        if c['nd']: ws.cell(row=r,column=13).number_format='dd mmm yyyy'
        if c['ceil']<5: ws.cell(row=r,column=7).fill=SHORT
        if not c['memok']: ws.cell(row=r,column=12).fill=SHORT
        if c.get('csp') and not c['csp'].lower().startswith('requirement met'):
            ws.cell(row=r,column=15).fill=SHORT
        for cc in range(1,len(head)+1): ws.cell(row=r,column=cc).border=THIN

    widths=[9,7,10,38,10,7,9,18,10,8,11,12,14,26,17]+[13]*12+[13,20]
    for i,w in enumerate(widths[:len(head)],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="E2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(head))}{ws.max_row}"

    # ---- area rollup ----
    a2=wb.create_sheet("By Area")
    a2.append(["Division","Area","Clubs","Avg goals met","Distinguished now",
               "Can still reach it","Cannot reach it","Short on next window","Meet membership rule"])
    for cc in range(1,10):
        cell=a2.cell(row=1,column=cc); cell.font=HDRF; cell.fill=HDRB
        cell.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center')
    a2.row_dimensions[1].height=46
    nxt=(L['agg'].get('close') or [{}])[0]
    grp=collections.defaultdict(list)
    for c in clubs: grp[(c['d'],c['a'])].append(c)
    for (d,a),g in sorted(grp.items(),key=lambda k:(k[0][0] or 'zz',k[0][1] or 'zz')):
        soon=sum(1 for c in g if c['nd']==nxt.get('date'))
        a2.append([d,a,len(g),round(sum(c['met'] for c in g)/len(g),2),
                   sum(1 for c in g if c['met']>=5),
                   sum(1 for c in g if c['met']<5 and c['ceil']>=5),
                   sum(1 for c in g if c['ceil']<5),soon,
                   sum(1 for c in g if c['memok'])])
    for i,w in enumerate([9,7,8,14,17,18,16,20,20],1): a2.column_dimensions[get_column_letter(i)].width=w
    a2.freeze_panes="C2"; a2.auto_filter.ref=a2.dimensions

    # ---- read me ----
    rm=wb.create_sheet("Read me",0)
    lines=[("District 21 - Distinguished Club Program, year in progress",True),
      ("",False),
      (f"Program year {L['py']}. Dashboard snapshot {L['asof']}. Built {L['generated']}.",False),
      (f"{L['days']} days remain until 30 June {L['py'][-4:]}, when these scores become final.",False),
      ("",False),
      ("Goals met counts 10 goals, not 12 rows",True),
      ("The club report prints twelve rows, but the DCP awards ten goals: the two officer-training",False),
      ("rows together earn one goal, and so do the two administrative rows. Counting rows instead of",False),
      ("goals overstates almost every club. The figures here follow the dashboard's own count.",False),
      ("",False),
      ("Ceiling",True),
      ("Goals already met plus those whose window is still open. It can only fall. A club whose",False),
      ("ceiling is below five can no longer be Distinguished this year, however well it finishes.",False),
      ("",False),
      ("Which windows close, and when",True)]
    for c in (L['agg'].get('close') or []):
        tail=f"{c['clubs']} clubs short" if c.get('open',True) else f"window opens {c['opens']}"
        lines.append((f"   {c['lbl']} - closes {c['date']} ({c['days']} days) - {tail}",False))
    lines+=[("",False),("Membership rule",True),
      ("Distinguished also requires 20 members or a net growth of five over base.",False),
      ("",False),("Colour",True),
      ("Green means the goal is met, pink means short, grey means the window has closed.",False),
      ("",False),
      ("Source: dashboards.toastmasters.org, read for each club individually.",False)]
    for t,b in lines:
        rm.append([t])
        if b: rm.cell(row=rm.max_row,column=1).font=BOLD
    rm.column_dimensions['A'].width=104

    p=_p('docs','inyear.xlsx'); wb.save(p)
    print(f"wrote {p}  {os.path.getsize(p)/1024:.0f} KB  clubs={len(clubs)} areas={len(grp)}")

if __name__=='__main__': main()
