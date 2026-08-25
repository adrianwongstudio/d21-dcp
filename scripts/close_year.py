"""Fold a finished program year into the published history.

The open year lives in live.json and is rebuilt from scratch every run. Once
a year ends on 30 June its scores are final, and it belongs with the other
finished years in data.json instead.

This does that incrementally: it scrapes only the year being closed and
merges it into the committed data.json, so it can run on a fresh checkout
where data/cache does not exist. Nothing else re-scrapes.

    python3 scripts/close_year.py              # close whatever is ready
    python3 scripts/close_year.py 2025-2026    # close a specific year
    python3 scripts/close_year.py --check      # say what would happen, change nothing

Exit codes: 0 did work or had nothing to do, 1 something went wrong.
"""
import os, sys, csv, io, gzip, json, queue, threading

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import common as C
import parse as P
from gen_site_data import year_record, transitions, MONTH_ORDER

THREADS = 8


def ready_to_close(today=None):
    """The most recent program year that has ended, if data.json lacks it."""
    today = today or C.today_local()
    # the year that ended this 30 June; before July there isn't one yet
    just_ended = f"{today.year - 1}-{today.year}"
    site = load_site()
    if just_ended in (site.get("years") or []):
        return None
    return just_ended


def archive_available(program_year):
    """Closed years get their own path; it 500s until the archive is built."""
    probe = C.load_clubs()[0][0]
    page = C.get(C.club_report_url(probe, program_year, 6, int(program_year[5:])))
    return bool(page and "<h2>" in page)


def load_site():
    with open(C.p("docs", "data.json"), encoding="utf-8") as fh:
        return json.load(fh)


def refresh_club_list(program_year):
    """Add any club that appeared in the closing year but is not in clubs.tsv."""
    ids = C.roster(program_year)
    if not ids:
        return 0
    raw = C.get(f"{C.BASE}/{program_year}/export.aspx"
                f"?type=CSV&report=clubperformance~{C.DISTRICT}~~~{program_year}")
    names = {}
    if raw:
        for row in csv.DictReader(io.StringIO(raw.lstrip("﻿"))):
            num = (row.get("Club Number") or "").strip()
            if num:
                names[num] = (row.get("Club Name") or "").strip()

    have = {cid for cid, _ in C.load_clubs()}
    added = [(n, names.get(n, f"Club {n}")) for n in sorted(ids) if n not in have]
    if not added:
        return 0
    with open(C.p("scripts", "clubs.tsv"), "a", encoding="utf-8") as fh:
        for num, name in added:
            fh.write(f"{num}\t{name}\n")
    return len(added)


def scrape_year(program_year):
    """Cache every club-month of one program year."""
    os.makedirs(C.CACHE, exist_ok=True)
    jobs = [(cid, m, y) for cid, _ in C.load_clubs()
            for (m, y) in C.months_of(program_year)]
    work = queue.Queue()
    for job in jobs:
        work.put(job)
    done = [0]
    lock = threading.Lock()

    def worker():
        while True:
            try:
                cid, m, y = work.get_nowait()
            except queue.Empty:
                return
            dest = os.path.join(C.CACHE, f"{cid}_{program_year}_{m:02d}.html.gz")
            if not (os.path.exists(dest) and os.path.getsize(dest) > 500):
                page = C.get(C.club_report_url(cid, program_year, m, y))
                if page:
                    with gzip.open(dest, "wt", encoding="utf-8") as fh:
                        fh.write(page)
            with lock:
                done[0] += 1
                if done[0] % 500 == 0:
                    print(f"  {done[0]}/{len(jobs)}", flush=True)

    threads = [threading.Thread(target=worker, daemon=True) for _ in range(THREADS)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    print(f"  scraped {done[0]}/{len(jobs)}", flush=True)


def parse_year(program_year):
    """{club id: (name, year record)} for the closing year."""
    from build import row_for
    out = {}
    for cid, name in C.load_clubs():
        months = {}
        for m, y in C.months_of(program_year):
            path = os.path.join(C.CACHE, f"{cid}_{program_year}_{m:02d}.html.gz")
            if not os.path.exists(path):
                continue
            try:
                with gzip.open(path, "rt", encoding="utf-8", errors="replace") as fh:
                    page = fh.read()
            except Exception:
                continue
            parsed = P.parse(page)
            if parsed.get("goals_met") is None and not parsed.get("goals"):
                continue
            months[m] = row_for(cid, name, program_year, m, y, parsed)
        if months:
            out[cid] = (name, year_record(months))
    return out


def merge(program_year, parsed):
    """Add the year to data.json and recompute the climbed/slipped lists."""
    site = load_site()
    if program_year in site["years"]:
        print(f"  {program_year} already present; nothing merged")
        return False

    by_id = {c["n"]: c for c in site["clubs"]}
    added_clubs = 0
    for cid, (name, record) in parsed.items():
        if record.get("f") is None and not any(record.get("s") or []):
            continue
        club = by_id.get(cid)
        if club is None:
            club = {"n": cid, "m": name, "d": record.get("d", ""),
                    "a": record.get("a", ""), "y": {}}
            by_id[cid] = club
            site["clubs"].append(club)
            added_clubs += 1
        club["y"][program_year] = record
        # the top-level alignment tracks the club's most recent year
        if record.get("d"):
            club["d"], club["a"] = record["d"], record["a"]

    site["years"] = sorted(set(site["years"]) | {program_year})
    site["clubs"].sort(key=lambda c: c["m"].lower())
    site["imp"], site["dec"] = transitions(site["clubs"], site["years"])
    site["generated"] = C.today_local().isoformat()

    with open(C.p("docs", "data.json"), "w", encoding="utf-8") as fh:
        json.dump(site, fh, separators=(",", ":"))

    scored = sum(1 for c in site["clubs"] if program_year in c["y"])
    print(f"  merged {program_year}: {scored} clubs "
          f"({added_clubs} new), years now {site['years']}")
    return True


def write_final_scores(program_year):
    """A one-sheet workbook of the closed year's final scores.

    This is the year-end handoff for the district spreadsheet: import it as a
    new sheet. Nothing here writes to Google Sheets — that step is manual.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    site = load_site()
    goals = site["goals"]
    head = (["Club No", "Club Name", "Division", "Area", "DCP Goals Met",
             "Status", "Club Success Plan", "Membership Base",
             "Members at Year End", "Net Growth"] + goals)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{program_year} final"
    ws.append(head)
    hdr, fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F3864")
    for i in range(1, len(head) + 1):
        c = ws.cell(row=1, column=i)
        c.font, c.fill = hdr, fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 42

    rows = 0
    for club in sorted(site["clubs"], key=lambda c: c["m"].lower()):
        y = club["y"].get(program_year)
        if not y or y.get("f") is None:
            continue
        net = (y["md"] - y["mb"]) if (y.get("md") is not None and y.get("mb") is not None) else None
        ws.append([club["n"], club["m"], y.get("d", ""), y.get("a", ""), y["f"],
                   y.get("st", ""), y.get("csp", ""), y.get("mb"), y.get("md"), net]
                  + list(y.get("g") or [None] * len(goals)))
        rows += 1

    widths = [11, 38, 9, 7, 9, 22, 20, 11, 12, 11] + [13] * len(goals)
    for i, w in enumerate(widths[:len(head)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(head))}{ws.max_row}"

    dest = C.p("output", f"{C.OUTPUT.get('inyear_prefix', 'District')}_{program_year}_final.xlsx")
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    wb.save(dest)
    print(f"  wrote {os.path.basename(dest)} ({rows} clubs) - import this into the spreadsheet")
    return dest


def widen_config(program_year):
    """Keep config.json's history window in step with what data.json holds."""
    with open(C.p("config.json"), encoding="utf-8") as fh:
        cfg = json.load(fh)
    start = int(cfg["history"]["start_year"])
    end = int(program_year[:4])
    want = end - start + 1
    if want > int(cfg["history"]["years"]):
        cfg["history"]["years"] = want
        with open(C.p("config.json"), "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, indent=2, ensure_ascii=False)
            fh.write("\n")
        print(f"  config history.years -> {want}")


def main(argv):
    check_only = "--check" in argv
    named = [a for a in argv if not a.startswith("-")]
    year = named[0] if named else ready_to_close()

    if not year:
        print("nothing to close: the most recent finished year is already published")
        return 0

    if year in (load_site().get("years") or []):
        print(f"nothing to close: {year} is already published")
        return 0

    if year == C.current_program_year():
        print(f"refusing to close {year}: it is still the open program year")
        return 1

    print(f"closing {year}")
    if not archive_available(year):
        print(f"  the dashboard has no archive for {year} yet; try again later")
        return 0

    if check_only:
        print("  --check: archive is ready, stopping without changing anything")
        return 0

    added = refresh_club_list(year)
    if added:
        print(f"  clubs.tsv +{added} clubs seen in {year}")
    scrape_year(year)
    parsed = parse_year(year)
    print(f"  parsed {len(parsed)} clubs")
    if not parsed:
        print("  nothing parsed; leaving data.json alone")
        return 1
    if merge(year, parsed):
        widen_config(year)
        write_final_scores(year)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
