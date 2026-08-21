import json,collections
import os as _os
_ROOT=_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
def _p(*a): return _os.path.join(_ROOT,*a)

from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter

GOALS=["Level 1 awards","Level 2 awards","More Level 2 awards","Level 3 awards",
"Level 4, Path Completion, or DTM award","One more Level 4, Path Completion, or DTM award",
"New members","More new members","Club officer roles trained June-August",
"Club officer roles trained November-February","Membership-renewal dues on time",
"Club officer list on time"]
PYS=["2021-2022","2022-2023","2023-2024","2024-2025","2025-2026"]
COLS=['Club No','Club Name','Division','Area','Program Year','Month','Month End','As Of',
      'DCP Status','Club Success Plan','Membership Base','Membership To Date','Net Growth',
      'DCP Goals Met']+GOALS

rows=json.load(open(_p('data','rows.json')))
rows.sort(key=lambda r:(r['Club Name'].lower(),r['Sort Key']))

HDR=Font(bold=True,color="FFFFFF"); FILL=PatternFill("solid",fgColor="1F3864")
def style(ws,ncols):
    for c in range(1,ncols+1):
        cell=ws.cell(row=1,column=c); cell.font=HDR; cell.fill=FILL
        cell.alignment=Alignment(vertical='center',wrap_text=True)
    ws.freeze_panes="C2"; ws.auto_filter.ref=ws.dimensions
    ws.row_dimensions[1].height=42

wb=Workbook(); wb.remove(wb.active)

# ---- Monthly Data ----
ws=wb.create_sheet("Monthly Data")
ws.append(COLS)
numeric={'Membership Base','Membership To Date','Net Growth','DCP Goals Met'}
for r in rows:
    out=[]
    for c in COLS:
        v=r.get(c)
        if c in numeric or c in GOALS:
            try: v=int(v)
            except (TypeError,ValueError): pass
        out.append(v)
    ws.append(out)
style(ws,len(COLS))
for i,c in enumerate(COLS,1):
    ws.column_dimensions[get_column_letter(i)].width = 34 if c=='Club Name' else (11 if c in('Club No','Month','Month End','As Of') else (22 if c=='DCP Status' else 15))

# ---- year-end finals ----
final={}
for r in rows:
    if r['Month'].startswith('Jun'):
        final[(r['Club No'],r['Program Year'])]={'dcp':r['DCP Goals Met'],'status':r['DCP Status'],
            'name':r['Club Name'],'mem':r['Membership To Date'],
            'div':r['Division'],'area':r['Area']}

names={r['Club No']:r['Club Name'] for r in rows}
ws=wb.create_sheet("Year-End Summary")
hdr=['Club No','Club Name','Division','Area']+[f"{p} DCP" for p in PYS]+[f"{p} Status" for p in PYS]
ws.append(hdr)
for cid in sorted(names,key=lambda c:names[c].lower()):
    last=next((final[(cid,p)] for p in reversed(PYS) if (cid,p) in final),{})
    row=[cid,names[cid],last.get('div',''),last.get('area','')]
    row+=[final.get((cid,p),{}).get('dcp') for p in PYS]
    row+=[final.get((cid,p),{}).get('status','') for p in PYS]
    ws.append(row)
style(ws,len(hdr))
ws.column_dimensions['B'].width=34
for i in range(3,len(hdr)+1): ws.column_dimensions[get_column_letter(i)].width=16

# ---- transitions ----
imp=[];dec=[]
for cid in names:
    for a,b in zip(PYS,PYS[1:]):
        pa=final.get((cid,a),{}).get('dcp'); pb=final.get((cid,b),{}).get('dcp')
        if pa is None or pb is None: continue
        if pa<5 and pb>pa:
            imp.append([cid,names[cid],final.get((cid,b),{}).get('div',''),final.get((cid,b),{}).get('area',''),a,pa,b,pb,pb-pa,final.get((cid,b),{}).get('status',''),
                        final.get((cid,a),{}).get('mem'),final.get((cid,b),{}).get('mem')])
        if pa>5 and pb<pa:
            dec.append([cid,names[cid],final.get((cid,b),{}).get('div',''),final.get((cid,b),{}).get('area',''),a,pa,b,pb,pb-pa,final.get((cid,b),{}).get('status',''),
                        final.get((cid,a),{}).get('mem'),final.get((cid,b),{}).get('mem')])
imp.sort(key=lambda r:-r[8]); dec.sort(key=lambda r:r[8])
th=['Club No','Club Name','Division','Area','From Year','From DCP','To Year','To DCP','Change',
    'To-Year Status','From-Year Members','To-Year Members']
for title,data,note in [("Improving Clubs",imp,"Finished a year under 5 DCP points, then improved the next year — success stories worth learning from."),
                        ("Declining Clubs",dec,"Finished a year above 5 DCP points, then dropped the next year — candidates for district support.")]:
    w=wb.create_sheet(title); w.append(th)
    for d in data: w.append(d)
    style(w,len(th)); w.column_dimensions['B'].width=34
    for i in range(3,len(th)+1): w.column_dimensions[get_column_letter(i)].width=17

# ---- Action List (most recent transition) ----
LA,LB=PYS[-2],PYS[-1]
w=wb.create_sheet("Action List 2025-2026")
ah=['Priority','Club No','Club Name','Division','Area',f'{LA} DCP',f'{LB} DCP','Change',
    f'{LB} Status',f'{LA} Members',f'{LB} Members']
w.append(ah)
recent_i=[x for x in imp if x[4]==LA and x[6]==LB]
recent_d=[x for x in dec if x[4]==LA and x[6]==LB]
for x in sorted(recent_d,key=lambda r:r[8]):
    w.append(['SUPPORT — declined from above 5',x[0],x[1],x[2],x[3],x[5],x[7],x[8],x[9],x[10],x[11]])
for x in sorted(recent_i,key=lambda r:-r[8]):
    w.append(['LEARN FROM — improved from under 5',x[0],x[1],x[2],x[3],x[5],x[7],x[8],x[9],x[10],x[11]])
style(w,len(ah)); w.column_dimensions['A'].width=34; w.column_dimensions['C'].width=34
for i in range(4,len(ah)+1): w.column_dimensions[get_column_letter(i)].width=15
RED=PatternFill("solid",fgColor="FCE4E4"); GRN=PatternFill("solid",fgColor="E2F0DA")
for row in w.iter_rows(min_row=2,max_row=w.max_row,max_col=len(ah)):
    f=RED if str(row[0].value).startswith('SUPPORT') else GRN
    for c in row: c.fill=f

# ---- README ----
w=wb.create_sheet("README",0)
info=[["District 21 — Club DCP Report"],[],
 ["Source","https://dashboards.toastmasters.org (ClubReport.aspx, District 21)"],
 ["Coverage",f"{len(names)} clubs x 60 month-ends ({PYS[0]} through {PYS[-1]})"],
 ["Rows in Monthly Data",len(rows)],
 ["Granularity","One row per club per month-end. Values are cumulative 'To Date' as of that month."],
 ["Note","Dashboard snapshots post ~1-2 weeks after month end; the 'As Of' column shows the actual snapshot date."],
 ["Note","DCP goal labels changed wording across years (e.g. 'Level 5' -> 'Path Completion'). Columns are aligned by DCP goal position 1-12."],
 ["Note","'DCP Status' is blank until a club qualifies as Distinguished or higher."],[],
 ["Tabs"],
 ["Monthly Data","Full month-by-month history. Use the filter row to slice by club, year or month."],
 ["Year-End Summary","Final (June 30) DCP score and status per club per year."],
 ["Improving Clubs","<5 DCP points, then improved the following year."],
 ["Declining Clubs",">5 DCP points, then declined the following year."],
 ["Action List 2025-2026","The most recent year's transitions only — the shortlist to act on now."]]
for r in info: w.append(r)
w['A1'].font=Font(bold=True,size=15); w.column_dimensions['A'].width=24; w.column_dimensions['B'].width=105
for c in ['A3','A4','A5','A6','A7','A8','A9','A12','A13','A14','A15','A16']: w[c].font=Font(bold=True)
w['A11'].font=Font(bold=True,size=12)
for cell in ['B6','B7','B8','B12','B13','B14','B15','B16']: w[cell].alignment=Alignment(wrap_text=True)

wb.save(_p("output","District21_DCP_Report.xlsx"))
print(f"rows={len(rows)} improving={len(imp)} declining={len(dec)}")

import csv
for nm,data in [("monthly_data",None),("improving_clubs",imp),("declining_clubs",dec)]:
    if data is None:
        with open(_p("output","monthly_data.csv"),"w",newline="") as f:
            wcsv=csv.DictWriter(f,fieldnames=COLS,extrasaction='ignore'); wcsv.writeheader()
            for r in rows: wcsv.writerow(r)
    else:
        with open(_p("output",nm+".csv"),"w",newline="") as f:
            wcsv=csv.writer(f); wcsv.writerow(th); wcsv.writerows(data)
